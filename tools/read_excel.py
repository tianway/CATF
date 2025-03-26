import openpyxl
from tools.logger import Logger
import os


class ReadExcel:

    logger = Logger.get_logger()

    @classmethod
    def read_excel_gui(cls, filepath, sheet_name, rs, re, c_no=1, c_title=2, c_data=3, c_expect=4):
        """
        读取gui测试数据excel文件中的内容
        :param filepath:测试数据excel文件路径
        :param sheet_name:sheet表名字
        :param rs:开始读取的行号
        :param re:结束读取的行号
        :param c_no:用例编号所在列号
        :param c_title:用例标题所在列号
        :param c_data:测试数据所在列号
        :param c_expect:期望结果所在列号
        :return:正常返回包含测试数据的字典对象，异常返回None
        """
        try:
            book = openpyxl.load_workbook(filepath)  # 获取book对象
            sheet = book[sheet_name]  # 获取sheet表对象
            testdata_dict = {}
            for i in range(rs, re):
                fuc_name = sheet.cell(i, 1).value.split("-")[0]
                if fuc_name not in testdata_dict.keys():
                    testdata_dict[fuc_name] = []
                temp = {}
                data_dict = {}
                if not sheet.cell(i, c_data).value is None:
                    data_list = sheet.cell(i, c_data).value.split("\n")
                    for j in data_list:
                        if j.split("=")[1] != 'null':
                            data_dict[j.split("=")[0]] = j.split("=")[1]
                        else:
                            data_dict[j.split("=")[0]] = ''
                temp["no"] = sheet.cell(i, c_no).value
                temp["title"] = sheet.cell(i, c_title).value
                temp["data"] = data_dict
                temp["expect"] = sheet.cell(i, c_expect).value
                testdata_dict[fuc_name].append(temp)
            cls.logger.info(f'{os.path.basename(filepath)}测试数据读取成功')
            return testdata_dict
        except:
            cls.logger.error(f'{os.path.basename(filepath)}测试数据读取异常')
            return None

    @classmethod
    def read_excel_api(cls, filepath, sheet_name, rs, re, c_no=1, c_title=2, c_method=3, c_uri=4, c_headers=5, c_data=6, c_expect=7):
        """
        读取api测试数据excel文件，返回字典对象
        :param filepath: 测试数据excel文件路径
        :param sheet_name: sheet表名字
        :param rs: 开始读取的行号
        :param re: 结束读取的行号
        :param c_no: 用例编号所在的列号
        :param c_title: 用例标题所在的列号
        :param c_method: 请求方法所在的列号
        :param c_uri: uri所在的列号
        :param c_headers: 请求头所在的列号
        :param c_data: 请求数据所在的列号
        :param c_expect: 期望结果所在的列号
        :return: 正常返回包含测试数据的字典对象，异常返回只包含键值的字典对象
        """
        try:
            book = openpyxl.load_workbook(filepath)  # 获取book对象
            sheet = book[sheet_name]  # 获取sheet表对象
            testdata_list = []
            for i in range(rs, re):
                temp = {}
                headers_dict = {}
                if sheet.cell(i, c_headers).value != 'null':
                    header_list = sheet.cell(i, c_headers).value.split("\n")
                    for line in header_list:
                        headers_dict[line.split("=")[0]] = line.split("=")[1]
                data_dict = {}
                if sheet.cell(i, c_data).value != 'null':
                    data_list = sheet.cell(i, c_data).value.split("\n")
                    for line in data_list:
                        data_dict[line.split("=")[0]] = line.split("=")[1]
                temp["no"] = sheet.cell(i, c_no).value
                temp["title"] = sheet.cell(i, c_title).value
                temp["method"] = sheet.cell(i, c_method).value
                temp["uri"] = sheet.cell(i, c_uri).value
                temp["headers"] = headers_dict
                temp["data"] = data_dict
                if sheet.cell(i, c_expect).value != 'null':
                    temp["expect"] = sheet.cell(i, c_expect).value
                else:
                    temp["expect"] = ''
                testdata_list.append(temp)
            cls.logger.info(f'{os.path.basename(filepath)}测试数据读取成功')
            return testdata_list
        except:
            cls.logger.error(f'{os.path.basename(filepath)}测试数据读取异常')
            return None


if __name__ == '__main__':
    sheet1 = ReadExcel.read_excel_gui('../data/testdata_gui.xlsx', '公共-系统版本', 2, 4)
    print(sheet1)
    sheet2 = ReadExcel.read_excel_api('../data/testdata_api.xlsx', '就业管理', 2, 33)
    print(sheet2)
